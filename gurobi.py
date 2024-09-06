# -*- coding: utf-8 -*-
"""
Distributed energy rosource network problem
Created: 14 nov 2023
@author: Alejandro Escudero-Santana (Code)
@author: Pablo Cortés-Achedad (Model)

Universidad de Sevilla
"""
import openpyxl
import pandas as pd
import scipy.spatial.distance as distance
from gurobipy import *
from itertools import permutations
import matplotlib.pyplot as plt
import numpy as np
import os
from pathlib import Path
from datetime import datetime


battery_path = Path(os.path.abspath(os.getcwd()), 'Battery_of_problems')
experiment = datetime.now().strftime("%Y%m%d_%H%M")
results_path = Path(os.path.abspath(os.getcwd()), 'Results')
os.makedirs(results_path, exist_ok=True)

name_file = 'BAT1_Escenario2.xlsx'
problem_path = Path(battery_path, name_file)


# Leer excel
wb = openpyxl.load_workbook(problem_path, data_only=True)
ws_description = wb['Description']
ws_data1 = wb['Data1']
ws_data2 = wb['Data2']

# Read size problem
size_J = ws_description['D2'].value
size_S = ws_description['D3'].value
size_K = ws_description['D4'].value
size_T = ws_description['D5'].value

# Read Data

# D_kt
D = {}
offset_column = 2
offset_row = 11
for t in range(size_T+1):
    for k in range(1, size_K+1):
        read_row = offset_row + t
        read_column = offset_column + k - 1
        D[k, t] = ws_data1.cell(row=read_row, column=read_column).value

# K_jkt
K = {}
offset_column = 2
offset_row = 6
for t in range(size_T+1):
    for j in range(1, size_J + 1):
        for k in range(1, size_K + 1):
            read_row = offset_row + t
            read_column = (k - 1) * size_J + j + offset_column - 1
            K[j, k, t] = ws_data2.cell(row=read_row, column=read_column).value

# IM_k
IM = {}
offset_column = 2
offset_row = 3
for column in range(size_K):
    read_row = offset_row
    read_column = offset_column + column
    IM[column + 1] = ws_data1.cell(row=read_row, column=read_column).value

# eta_jk
eta = {}
offset_column = 2
offset_row = 4
for j in range(1, size_J+1):
    for k in range(1, size_K+1):
        read_row = offset_row
        read_column = (k-1)*size_J + j + offset_column - 1
        eta[j, k] = ws_data2.cell(row=read_row, column=read_column).value

# ro_k
ro = {}
offset_column = 2
offset_row = 5
for column in range(size_K):
    read_row = offset_row
    read_column = offset_column + column
    ro[column + 1] = ws_data1.cell(row=read_row, column=read_column).value

# F_j
F = {}
offset_column = 3 + size_K
offset_row = 9
for column in range(size_J):
    read_row = offset_row
    read_column = offset_column + column
    F[column + 1] = ws_data1.cell(row=read_row, column=read_column).value

# C_jt
C = {}
offset_column = 3 + size_K
offset_row = 11
for t in range(size_T+1):
    for j in range(1, size_J+1):
        read_row = offset_row + t
        read_column = offset_column + j - 1
        C[j, t] = ws_data1.cell(row=read_row, column=read_column).value


# S_k
S = {}
offset_column = 2
offset_row = 7
for column in range(size_K):
    read_row = offset_row
    read_column = offset_column + column
    S[column + 1] = ws_data1.cell(row=read_row, column=read_column).value

# V_k
V = {}
offset_column = 2
offset_row = 9
for k in range(1, size_K+1):
    read_row = offset_row
    read_column = offset_column + k - 1
    V[k] = ws_data1.cell(row=read_row, column=read_column).value

# P_t
P = {}
offset_column = 4 + size_K + size_J
offset_row = 11
for row in range(size_T+1):
    read_row = offset_row + row
    read_column = offset_column
    P[row] = ws_data1.cell(row=read_row, column=read_column).value

try:
    # Create a new model
    m = Model('DERNP')
    m.params.LogFile = './logfile.txt'

    # Create sets
    set_J = list(range(1, size_J+1))  # Set of energies generation sources
    set_S = [2, 3]  # Set of thermal energy generation
    set_K = [1, 2]  # Set of the commodities: electricity (1) and heating (2)
    set_T = list(range(size_T+1))  # Temporal horizont

    # Create variable
    x = {}  # x_jkt: Flow of generated energy of type k by source j in period t
    I = {}  # l_kt: Level of stored energy of type k at the end of period t
    u = {}  # u_kt: Amount of energy that is charged at the storage facility of type k energy in the period t
    v = {}  # v_kt: Amount of energy that is discharged from the storage facility of type k energy to be supplied to
    # the building installation in the period
    g = {}  # g_jt: Amount of gas acquired to the gas grid provider with destination the energy generation source j
    # (corresponding to the boiler or the CHP) in period
    e = {}  # e_kt: Excess of energy of type k that is generated in period t to be sold (in case of electricity) or
    # to be wasted in case of heating (that should be zero in case of optimality)
    y = {}  # y_jkt: It takes vale equal to 1 in case of the energy generation source j is activated to supply type k
    # energy in period t, and 0 otherwise
    a = {}  # a_jkt: Continuous variable that takes a vale equal to 1 in case of the energy generation source j is
    # initialised to supply type k energy in period t, and 0 otherwise.
    z = {}  # z_kt: Binary variable. It takes vale equal to 1 in case of the storage facility of type k energy is
    # activated in period t, and 0 otherwise.
    b = {} # b_kt: Continuous variable that takes a vale equal to 1 in case of the storage facility of type k energy is
    # initialised in period t, and 0 otherwise

    for j in set_J:
        for k in set_K:
            for t in set_T:
                x[j, k, t] = m.addVar(vtype=GRB.CONTINUOUS, lb=0.0, name='x_%s_%s_%s' % (j, k, t))

    for k in set_K:
        for t in set_T:
            I[k, t] = m.addVar(vtype=GRB.CONTINUOUS, lb=0.0, name='I_%s_%s' % (k, t))

    for k in set_K:
        for t in set_T:
            u[k, t] = m.addVar(vtype=GRB.CONTINUOUS, lb=0.0, name='u_%s_%s' % (k, t))

    for k in set_K:
        for t in set_T:
            v[k, t] = m.addVar(vtype=GRB.CONTINUOUS, lb=0.0, name='v_%s_%s' % (k, t))

    for j in set_J:
        for t in set_T:
            g[j, t] = m.addVar(vtype=GRB.CONTINUOUS, lb=0.0, name='g_%s_%s' % (j, t))

    for k in set_K:
        for t in set_T:
            e[k, t] = m.addVar(vtype=GRB.CONTINUOUS, lb=0.0, name='e_%s_%s' % (k, t))

    for j in set_J:
        for k in set_K:
            for t in set_T:
                #if j != 1:
                y[j, k, t] = m.addVar(vtype=GRB.BINARY, name='y_%s_%s_%s' % (j, k, t))

    for j in set_J:
        for k in set_K:
            for t in set_T:
                #if j != 1:
                a[j, k, t] = m.addVar(vtype=GRB.CONTINUOUS, lb=0.0, name='a_%s_%s_%s' % (j, k, t))

    for k in set_K:
        for t in set_T:
            z[k, t] = m.addVar(vtype=GRB.BINARY, name='z_%s_%s' % (k, t))

    for k in set_K:
        for t in set_T:
            b[k, t] = m.addVar(vtype=GRB.CONTINUOUS, lb=0.0, name='b_%s_%s' % (k, t))




    # Integrate new variables
    m.update()
    # Set objective
    objectiveFunction = LinExpr()
    for t in set_T:
        for k in set_K:
            for j in set_J:
                if j != 1:
                    objectiveFunction.addTerms(F[j], a[j, k, t])
            for j in set_J:
                if j not in set_S:
                    objectiveFunction.addTerms(C[j, t], x[j, k, t])

            for j in set_S:
                if k == 2:  # Add to sum g only one time
                    objectiveFunction.addTerms(C[j, t], g[j, t])

            objectiveFunction.addTerms(S[k], b[k, t])

            objectiveFunction.addTerms(V[k], u[k, t])

            objectiveFunction.addTerms(V[k], v[k, t])

        objectiveFunction.addTerms(-P[t], e[1, t])
        objectiveFunction.addTerms(0.1, e[2, t]) #TODO Añadido para penalizar excedente innecesario

    m.setObjective(objectiveFunction, GRB.MINIMIZE)

    # Constraint (2)
    ind = 0
    for k in set_K:
        for t in set_T[1:]:
            ind += 1
            m.addConstr(I[k, t-1] + quicksum(x[j, k, t] for j in set_J) == D[k, t] + e[k, t] + I[k, t], 'C2.%d' % ind)

    # Constraint (3)
    ind = 0
    for k in set_K:
        for t in set_T[1:]:
            ind += 1
            m.addConstr(I[k, t] - I[k, t-1] == u[k, t] - v[k, t], 'C3.%d' % ind)

    # Constraint (4)
    ind = 0
    for k in set_K:
        for t in set_T[1:]:
            ind += 1
            #m.addConstr(u[k, t] >= ro[k] * quicksum(x[j, k, t] - D[k, t] - e[k, t] for j in set_J), 'C4.%d' % ind)
            m.addConstr(u[k, t] >= ro[k] * (quicksum(x[j, k, t] for j in set_J) - D[k, t] - e[k, t]), 'C4.%d' % ind)
    # Constraint (5)
    ind = 0
    for k in set_K:
        for t in set_T[1:]:
            ind += 1
            m.addConstr(v[k, t] >= D[k, t] + e[k, t] - quicksum(x[j, k, t] for j in set_J), 'C5.%d' % ind)

    # Constraint (6)
    ind = 0
    for k in set_K:
        for t in set_T[1:]:
            ind += 1
            m.addConstr(u[k, t] + v[k, t] <= IM[k] * z[k, t], 'C6.%d' % ind)

    # Constraint (7)
    ind = 0
    for k in set_K:
        for t in set_T[1:]:
            ind += 1
            m.addConstr(b[k, t] >= z[k, t] - z[k, t-1] , 'C7.%d' % ind)

    # Constraint (8)
    ind = 0
    for k in set_K:
        for t in set_T[1:]:
            ind += 1
            m.addConstr(I[k, t] <= IM[k], 'C8.%d' % ind)

    # Constraint (9)
    ind = 0
    for j in set_J:
        for k in set_K:
            for t in set_T[1:]:
                #if j != 1:
                ind += 1
                m.addConstr(x[j, k, t] <= K[j, k, t] * y[j, k, t], 'C9.%d' % ind)

    # Constraint (10)
    ind = 0
    for j in set_J:
        for k in set_K:
            for t in set_T[1:]:
                #if j != 1:
                ind += 1
                m.addConstr(a[j, k, t] >= y[j, k, t] - y[j, k, t-1], 'C10.%d' % ind)

    # Constraint (11)
    ind = 0
    for j in set_S:
        for k in set_K:
            for t in set_T[1:]:
                ind += 1
                m.addConstr(x[j, k, t] == eta[j, k] * g[j, t], 'C11.%d' % ind)

    # Constraint (15)
    ind = 0
    for k in set_K:
        ind += 1
        #m.addConstr(I[k, 0] == 0, 'C15A.%d' % ind)
        m.addConstr(I[k, 0] == 1000, 'C15A1.%d' % ind)
        #m.addConstr(e[k, 1] == 0, 'C15A2.%d' % ind)

    ind = 0
    for j in set_J:
        for k in set_K:
            ind += 1
            m.addConstr(y[j, k, 0] == 0, 'C15B.%d' % ind)

    ind = 0
    for k in set_K:
        ind += 1
        #m.addConstr(z[k, 0] == 0, 'C15C.%d' % ind)
        m.addConstr(z[k, 0] == 1, 'C15C.%d' % ind)


    m.params.timeLimit = 1000.0
    # m.params.presolve = 0
    # m.params.tolerance = 10 ** -9
    m.params.MIPGap = 0

    # Optimize the model
    # m.write("file.lp")
    m.optimize()

    m.write("model.lp")
    # print(m.Runtime)
    # print(m.Runtime)
    # print(self.model.Runtime)
    # for var in m.getVars():
    #    if var.x != 0:
    #        print(var.varName, var.x)

    print('Obj:', m.objVal)

except GurobiError:
    print('Error Gurobi')

# Variables
list_variable = m.getVars()

